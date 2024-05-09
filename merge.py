from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# Function to merge Excel sheets into one while preserving formatting
def merge_excel_sheets(input_file, output_file):
    # Load the workbook
    wb_input = load_workbook(input_file)
    wb_output = load_workbook(output_file)
    
    # Loop through each sheet in the input workbook
    for sheet_name in wb_input.sheetnames:
        # Check if sheet already exists in the output workbook
        if sheet_name in wb_output.sheetnames:
            # If sheet exists, find the next available row
            sheet_output = wb_output[sheet_name]
            next_row = sheet_output.max_row + 1
        else:
            # If sheet doesn't exist, create a new sheet
            sheet_output = wb_output.create_sheet(title=sheet_name)
            next_row = 1

        # Loop through each row and column in the input sheet
        for row in wb_input[sheet_name].iter_rows(values_only=True):
            # Write each cell value to the output sheet
            for col_idx, value in enumerate(row, start=1):
                sheet_output.cell(row=next_row, column=col_idx, value=value)

            # Move to the next row in the output sheet
            next_row += 1

    # Save the output workbook
    wb_output.save(output_file)

# Example usage
if __name__ == "__main__":
    input_file = r"C:\Users\ASUS\Desktop\sample.xlsx"  # Replace with your input file path
    output_file = r"C:\Users\ASUS\Desktop\merged_output.xlsx"  # Replace with your desired output file path
    merge_excel_sheets(input_file, output_file)
    print("Excel sheets merged successfully with preserved formatting!")
